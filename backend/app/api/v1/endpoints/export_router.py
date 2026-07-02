import csv
import io
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime

from app.database.session import get_db
from app.models.timetable import Timetable, TimeSlot, RoomAllocation
from app.models.section import Section
from app.models.faculty import Faculty, FacultyWorkload
from app.models.room import Classroom, Laboratory
from app.models.subject import Course
from app.services.conflict_checker import validate_timetable_entry

# ReportLab imports
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Openpyxl imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter()

# ==========================================
# HELPERS FOR DATA GATHERING
# ==========================================

def get_timetable_data(db: Session, filters: dict) -> list:
    query = db.query(Timetable).filter(Timetable.academic_year == filters.get("academic_year", "2025-2026"))
    if filters.get("section_id"):
        query = query.filter(Timetable.section_id == filters["section_id"])
    if filters.get("faculty_id"):
        query = query.filter(Timetable.faculty_id == filters["faculty_id"])
    if filters.get("classroom_id"):
        query = query.join(RoomAllocation).filter(RoomAllocation.classroom_id == filters["classroom_id"])
    if filters.get("lab_id"):
        query = query.join(RoomAllocation).filter(RoomAllocation.lab_id == filters["lab_id"])
        
    results = query.all()
    
    data = []
    for item in results:
        ra = item.room_allocation
        room_num = "TBD"
        if ra:
            if ra.room_type == "Classroom" and ra.classroom:
                room_num = ra.classroom.room_number
            elif ra.room_type == "Laboratory" and ra.lab:
                room_num = ra.lab.lab_number
                
        data.append({
            "day": item.slot.day_of_week,
            "period": item.slot.slot_order,
            "time": f"{item.slot.start_time} - {item.slot.end_time}",
            "course": f"{item.course.course_code} - {item.course.course_name}",
            "course_type": item.course.course_type,
            "faculty": f"{item.faculty.first_name} {item.faculty.last_name}",
            "section": item.section.section_name,
            "room": room_num
        })
    return data

# ==========================================
# TIMETABLE EXPORTS
# ==========================================

@router.get("/timetable/csv")
def export_timetable_csv(
    section_id: Optional[int] = Query(None),
    faculty_id: Optional[int] = Query(None),
    classroom_id: Optional[int] = Query(None),
    lab_id: Optional[int] = Query(None),
    academic_year: str = Query("2025-2026"),
    db: Session = Depends(get_db)
):
    filters = {
        "section_id": section_id,
        "faculty_id": faculty_id,
        "classroom_id": classroom_id,
        "lab_id": lab_id,
        "academic_year": academic_year
    }
    records = get_timetable_data(db, filters)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(["Day", "Period", "Time Slot", "Course", "Type", "Faculty", "Section", "Room Number"])
    
    for r in records:
        writer.writerow([r["day"], r["period"], r["time"], r["course"], r["course_type"], r["faculty"], r["section"], r["room"]])
        
    output.seek(0)
    response = StreamingResponse(iter([output.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename=timetable_export_{academic_year}.csv"
    return response

@router.get("/timetable/excel")
def export_timetable_excel(
    section_id: Optional[int] = Query(None),
    faculty_id: Optional[int] = Query(None),
    classroom_id: Optional[int] = Query(None),
    lab_id: Optional[int] = Query(None),
    academic_year: str = Query("2025-2026"),
    db: Session = Depends(get_db)
):
    filters = {
        "section_id": section_id,
        "faculty_id": faculty_id,
        "classroom_id": classroom_id,
        "lab_id": lab_id,
        "academic_year": academic_year
    }
    records = get_timetable_data(db, filters)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Timetable"
    
    # Design Styles
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1E3A8A")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=11)
    
    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border_side = Side(style='thin', color='E2E8F0')
    border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    # Write Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"COLLEGE TIMETABLE ALLOCATION - AY {academic_year}"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 40
    
    # Write headers
    headers = ["Day", "Period", "Time Slot", "Course Name / Code", "Type", "Faculty Member", "Section", "Room Number"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_data
    ws.row_dimensions[3].height = 28
    
    # Write data
    row_idx = 4
    for r in records:
        row_data = [r["day"], r["period"], r["time"], r["course"], r["course_type"], r["faculty"], r["section"], r["room"]]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_data
            cell.alignment = align_center if col_idx in [1, 2, 3, 5, 7, 8] else align_left
            if row_idx % 2 == 0:
                cell.fill = fill_alt
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1
        
    # Auto-fit columns width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = StreamingResponse(iter([output.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename=timetable_export_{academic_year}.xlsx"
    return response

@router.get("/timetable/pdf")
def export_timetable_pdf(
    section_id: Optional[int] = Query(None),
    faculty_id: Optional[int] = Query(None),
    classroom_id: Optional[int] = Query(None),
    lab_id: Optional[int] = Query(None),
    academic_year: str = Query("2025-2026"),
    db: Session = Depends(get_db)
):
    filters = {
        "section_id": section_id,
        "faculty_id": faculty_id,
        "classroom_id": classroom_id,
        "lab_id": lab_id,
        "academic_year": academic_year
    }
    records = get_timetable_data(db, filters)
    
    # Set PDF parameters (Landscape for wide columns grid)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor("#1E3A8A"),
        spaceAfter=15,
        alignment=1 # Center
    )
    
    cell_header_style = ParagraphStyle(
        'HeaderStyle',
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1
    )
    
    cell_data_style = ParagraphStyle(
        'DataStyle',
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor("#334155")
    )
    
    # Title Paragraph
    title_text = f"Weekly Timetable Schedule Grid - Academic Year {academic_year}"
    if section_id:
        sec = db.query(Section.section_name).filter(Section.section_id == section_id).scalar()
        title_text += f" (Section: {sec})"
    elif faculty_id:
        fac = db.query(Faculty).filter(Faculty.faculty_id == faculty_id).first()
        if fac:
            title_text += f" (Faculty: {fac.first_name} {fac.last_name})"
            
    story.append(Paragraph(title_text, title_style))
    story.append(Spacer(1, 10))
    
    # Format Table Data
    table_content = [[
        Paragraph("Day", cell_header_style),
        Paragraph("Period", cell_header_style),
        Paragraph("Time Slot", cell_header_style),
        Paragraph("Course", cell_header_style),
        Paragraph("Type", cell_header_style),
        Paragraph("Faculty Member", cell_header_style),
        Paragraph("Section", cell_header_style),
        Paragraph("Room", cell_header_style)
    ]]
    
    for r in records:
        table_content.append([
            Paragraph(r["day"], cell_data_style),
            Paragraph(str(r["period"]), cell_data_style),
            Paragraph(r["time"], cell_data_style),
            Paragraph(r["course"], cell_data_style),
            Paragraph(r["course_type"], cell_data_style),
            Paragraph(r["faculty"], cell_data_style),
            Paragraph(r["section"], cell_data_style),
            Paragraph(r["room"], cell_data_style)
        ])
        
    # Table Styling
    col_widths = [70, 45, 95, 170, 50, 110, 55, 55] # sum is 650 (fit inside margins of landscape letter 792 width)
    t = Table(table_content, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1E3A8A")),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
        ('TOPPADDING', (0,1), (-1,-1), 6),
    ]))
    
    story.append(t)
    doc.build(story)
    
    buffer.seek(0)
    response = StreamingResponse(iter([buffer.getvalue()]), media_type="application/pdf")
    response.headers["Content-Disposition"] = f"attachment; filename=timetable_report_{academic_year}.pdf"
    return response

# ==========================================
# FACULTY WORKLOAD EXPORTS
# ==========================================

def get_faculty_report_data(db: Session, academic_year: str) -> list:
    faculty_list = db.query(Faculty).all()
    report = []
    designation_limits = {
        "Professor": 12,
        "Associate Professor": 14,
        "Assistant Professor": 16,
        "Lecturer": 18
    }
    for f in faculty_list:
        max_allowed = designation_limits.get(f.designation, 18)
        scheduled_hours = db.query(Timetable).filter(
            Timetable.faculty_id == f.faculty_id,
            Timetable.academic_year == academic_year
        ).count()
        utilization = round((scheduled_hours / max_allowed) * 100, 1) if max_allowed > 0 else 0
        report.append({
            "code": f.employee_code,
            "name": f"{f.first_name} {f.last_name}",
            "designation": f.designation,
            "scheduled": scheduled_hours,
            "limit": max_allowed,
            "utilization": utilization
        })
    return report

@router.get("/faculty/csv")
def export_faculty_workload_csv(
    academic_year: str = Query("2025-2026"),
    db: Session = Depends(get_db)
):
    records = get_faculty_report_data(db, academic_year)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Emp Code", "Faculty Name", "Designation", "Hours Scheduled", "Hours Limit", "Utilization Percentage"])
    for r in records:
        writer.writerow([r["code"], r["name"], r["designation"], r["scheduled"], r["limit"], f"{r['utilization']}%"])
    output.seek(0)
    response = StreamingResponse(iter([output.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename=faculty_workload_{academic_year}.csv"
    return response

@router.get("/faculty/excel")
def export_faculty_workload_excel(
    academic_year: str = Query("2025-2026"),
    db: Session = Depends(get_db)
):
    records = get_faculty_report_data(db, academic_year)
    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty Workloads"
    
    font_title = Font(name="Segoe UI", size=14, bold=True, color="1E3A8A")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=11)
    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'), top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))
    
    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"FACULTY WORKLOAD UTILIZATION REPORT - AY {academic_year}"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # Headers
    headers = ["Emp Code", "Faculty Name", "Designation", "Hours Scheduled", "Hours Limit", "Utilization %"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[3].height = 25
    
    row_idx = 4
    for r in records:
        row_data = [r["code"], r["name"], r["designation"], r["scheduled"], r["limit"], f"{r['utilization']}%"]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 4, 5, 6] else "left", vertical="center")
            if row_idx % 2 == 0:
                cell.fill = fill_alt
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = StreamingResponse(iter([output.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename=faculty_workload_{academic_year}.xlsx"
    return response

@router.get("/faculty/pdf")
def export_faculty_workload_pdf(
    academic_year: str = Query("2025-2026"),
    db: Session = Depends(get_db)
):
    records = get_faculty_report_data(db, academic_year)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor("#1E3A8A"),
        spaceAfter=15,
        alignment=1
    )
    cell_header = ParagraphStyle('Head', fontName='Helvetica-Bold', fontSize=10, textColor=colors.white, alignment=1)
    cell_data = ParagraphStyle('Data', fontName='Helvetica', fontSize=9, textColor=colors.HexColor("#334155"))
    
    story.append(Paragraph(f"Faculty Workload Report - AY {academic_year}", title_style))
    story.append(Spacer(1, 10))
    
    table_content = [[
        Paragraph("Emp Code", cell_header),
        Paragraph("Faculty Name", cell_header),
        Paragraph("Designation", cell_header),
        Paragraph("Scheduled", cell_header),
        Paragraph("Limit", cell_header),
        Paragraph("Utilization %", cell_header)
    ]]
    for r in records:
        table_content.append([
            Paragraph(r["code"], cell_data),
            Paragraph(r["name"], cell_data),
            Paragraph(r["designation"], cell_data),
            Paragraph(f"{r['scheduled']} hrs", cell_data),
            Paragraph(f"{r['limit']} hrs", cell_data),
            Paragraph(f"{r['utilization']}%", cell_data)
        ])
        
    t = Table(table_content, colWidths=[70, 140, 130, 65, 55, 75])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1E3A8A")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('ALIGN', (3,0), (-1,-1), 'CENTER')
    ]))
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    response = StreamingResponse(iter([buffer.getvalue()]), media_type="application/pdf")
    response.headers["Content-Disposition"] = f"attachment; filename=faculty_workload_{academic_year}.pdf"
    return response

# ==========================================
# ROOM UTILIZATION EXPORTS
# ==========================================

def get_room_report_data(db: Session, academic_year: str) -> list:
    classrooms = db.query(Classroom).filter(Classroom.is_available == True).all()
    laboratories = db.query(Laboratory).filter(Laboratory.is_available == True).all()
    total_slots = 40.0
    report = []
    
    for c in classrooms:
        booked = db.query(RoomAllocation).join(Timetable).filter(
            RoomAllocation.classroom_id == c.classroom_id,
            RoomAllocation.room_type == "Classroom",
            Timetable.academic_year == academic_year
        ).count()
        util = round((booked / total_slots) * 100, 1)
        report.append({
            "number": c.room_number,
            "type": "Classroom",
            "building": c.building,
            "capacity": c.capacity,
            "booked": booked,
            "utilization": util
        })
        
    for l in laboratories:
        booked = db.query(RoomAllocation).join(Timetable).filter(
            RoomAllocation.lab_id == l.lab_id,
            RoomAllocation.room_type == "Laboratory",
            Timetable.academic_year == academic_year
        ).count()
        util = round((booked / total_slots) * 100, 1)
        report.append({
            "number": l.lab_number,
            "type": "Laboratory",
            "building": l.building,
            "capacity": l.capacity,
            "booked": booked,
            "utilization": util
        })
    return report

@router.get("/rooms/csv")
def export_room_utilization_csv(
    academic_year: str = Query("2025-2026"),
    db: Session = Depends(get_db)
):
    records = get_room_report_data(db, academic_year)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Room Number", "Room Type", "Building", "Capacity", "Booked Slots", "Utilization %"])
    for r in records:
        writer.writerow([r["number"], r["type"], r["building"], r["capacity"], r["booked"], f"{r['utilization']}%"])
    output.seek(0)
    response = StreamingResponse(iter([output.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename=room_utilization_{academic_year}.csv"
    return response

@router.get("/rooms/excel")
def export_room_utilization_excel(
    academic_year: str = Query("2025-2026"),
    db: Session = Depends(get_db)
):
    records = get_room_report_data(db, academic_year)
    wb = Workbook()
    ws = wb.active
    ws.title = "Room Utilization"
    
    font_title = Font(name="Segoe UI", size=14, bold=True, color="1E3A8A")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=11)
    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'), top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))
    
    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"CLASSROOM & LAB OCCUPANCY UTILIZATION - AY {academic_year}"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # Headers
    headers = ["Room Number", "Room Type", "Building", "Capacity", "Booked Slots", "Utilization %"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[3].height = 25
    
    row_idx = 4
    for r in records:
        row_data = [r["number"], r["type"], r["building"], r["capacity"], r["booked"], f"{r['utilization']}%"]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 2, 4, 5, 6] else "left", vertical="center")
            if row_idx % 2 == 0:
                cell.fill = fill_alt
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = StreamingResponse(iter([output.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename=room_utilization_{academic_year}.xlsx"
    return response

@router.get("/rooms/pdf")
def export_room_utilization_pdf(
    academic_year: str = Query("2025-2026"),
    db: Session = Depends(get_db)
):
    records = get_room_report_data(db, academic_year)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor("#1E3A8A"),
        spaceAfter=15,
        alignment=1
    )
    cell_header = ParagraphStyle('Head', fontName='Helvetica-Bold', fontSize=10, textColor=colors.white, alignment=1)
    cell_data = ParagraphStyle('Data', fontName='Helvetica', fontSize=9, textColor=colors.HexColor("#334155"))
    
    story.append(Paragraph(f"Classroom & Lab Occupancy Report - AY {academic_year}", title_style))
    story.append(Spacer(1, 10))
    
    table_content = [[
        Paragraph("Room Number", cell_header),
        Paragraph("Room Type", cell_header),
        Paragraph("Building Name", cell_header),
        Paragraph("Capacity", cell_header),
        Paragraph("Booked Slots", cell_header),
        Paragraph("Utilization %", cell_header)
    ]]
    for r in records:
        table_content.append([
            Paragraph(r["number"], cell_data),
            Paragraph(r["type"], cell_data),
            Paragraph(r["building"], cell_data),
            Paragraph(f"{r['capacity']} seats", cell_data),
            Paragraph(f"{r['booked']} / 40", cell_data),
            Paragraph(f"{r['utilization']}%", cell_data)
        ])
        
    t = Table(table_content, colWidths=[90, 95, 125, 75, 75, 75])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1E3A8A")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('ALIGN', (3,0), (-1,-1), 'CENTER')
    ]))
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    response = StreamingResponse(iter([buffer.getvalue()]), media_type="application/pdf")
    response.headers["Content-Disposition"] = f"attachment; filename=room_utilization_{academic_year}.pdf"
    return response
